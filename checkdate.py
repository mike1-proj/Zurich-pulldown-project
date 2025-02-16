"""
this is a function opening remote Excel file and then checking a particular cell value on a particular
sheet. This can be used to make sure not to duplicate data or overwriting the same data if the same dated data
already exists in the work book sheet.


"""
from datetime import date

from openpyxl import load_workbook


def check_date():
    # Defining the path which Excel needs to be created
    filepath = "/home/michael/Desktop/new fund mix anlysis Zurich24.xlsx"
    # Generating workbook instance with our existing Excel file as the target
    excel_workbook = load_workbook(filepath, keep_vba=True)
    ab = excel_workbook
    ts = ab["analysisnew"]  # this makes variable ts equal the sheet tab we want
    bs = ts["B3"].value  # this gives us the date contained in the Excel sheet "analysisnew" cell B3
    # this  next line was used to get today's date but is not now required for this particular version
    # date1 = (date.today().strftime("%d/%m/%y"))  # format today's date so it matches Excel sheet format
    result = bs
    return result


def read_diff_value():
    # this function looks for the fund performance difference value in analysisnew tab
    # note if formula is returned instead of value open sheet go to cell and press F9 and save and close file
    # make sure to include data_only arg so only value is seen when sheet is opened and cell value is checked
    filepath = "/home/michael/Desktop/new fund mix anlysis Zurich24.xlsx"
    # Generating workbook instance with our existing Excel file as the target
    excel_workbook = load_workbook(filepath, keep_vba=True, data_only=True)
    ab = excel_workbook
    ts = ab["analysisnew"]  # this makes variable ts equal the sheet tab we want
    bs = ts["C20"].value  # this gives us the value contained in the Excel sheet "analysisnew" cell C20
    result = bs
    return result








